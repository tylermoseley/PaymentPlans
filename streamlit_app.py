import streamlit as st
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
import tempfile
import warnings
import copy

if __name__ == '__main__':

    # Header and uploads
    st.set_page_config(layout="wide")
    st.title('Payment Plan Update Tool')

    uploaded_PayPlan = st.file_uploader("Upload Haemonetics Payment Plan Matrix", type=['xlsx', 'xlsm'])

    if uploaded_PayPlan is not None:
        with warnings.catch_warnings(record=True):
            warnings.simplefilter("ignore", category=UserWarning)
            PayPlans = pd.read_excel(io=uploaded_PayPlan, header=11)

    uploaded_MarketingUpdate = st.file_uploader("Upload Marketing Update File", type=['xlsx', 'xlsm'])

    if uploaded_MarketingUpdate:
        # MarketingUpdates = pd.read_excel(uploaded_MarketingUpdate, header=11)
        MarketingUpdates = pd.read_excel(uploaded_MarketingUpdate)
        Mappings = pd.read_excel(io=uploaded_MarketingUpdate, sheet_name='Lookups', usecols='E:I', header=0)
        Specialty = pd.read_excel(io=uploaded_MarketingUpdate,sheet_name='Specialty',usecols='B:J',header=2)
        ProposeCols = [None] + [col for col in MarketingUpdates.columns if col.lower().startswith('proposal')]
        ProposeCol = st.selectbox('Proposal To Load:', ProposeCols)

        if ProposeCol:
            # Transformations
            OriginalPayPlans = PayPlans.copy()
            OriginalPayPlans.set_index('Payment Ruleset Code:Name', inplace=True)
            PayPlans['RULENAME'] = PayPlans['Payment Ruleset Code:Name'].str.split(':').str[1]

            # Primary Payment updates
            MarketingUpdatesMod = (MarketingUpdates
                .loc[MarketingUpdates['#'].isna()]
                .loc[MarketingUpdates['IT Use.1'].notna()]
                .loc[MarketingUpdates['DONOR COMPENSATION'].notna()]
                .merge(Mappings, how='left', left_on='DONOR COMPENSATION', right_on='Business Name')
                .merge(PayPlans, how='left', left_on='RULENAME', right_on='RULENAME')
                [['IT Use.1', 'Payment Ruleset Code:Name', ProposeCol, 'Business Name']]
                )
            # Pivot
            MarketingUpdatesMod = MarketingUpdatesMod.pivot_table(index='Payment Ruleset Code:Name', columns='IT Use.1',
                                                                  values=ProposeCol)
            # Append "Amount ($): " to column headers for amounts
            MarketingUpdatesMod.columns = ["Amount ($): " + col for col in MarketingUpdatesMod.columns]

            # Specialty Payment updates
            SpecialtyMod = (Specialty
            .merge(PayPlans, how='left', left_on='Name (calculated)', right_on='RULENAME')
            [['PROFILECODE', 'Payment Ruleset Code:Name', 'new Fee']]
            )
            SpecialtyMod = SpecialtyMod.pivot_table(index='Payment Ruleset Code:Name', columns='PROFILECODE',
                                                    values='new Fee')
            SpecialtyMod.columns = ["Amount ($): " + col for col in SpecialtyMod.columns]

            pd.set_option('future.no_silent_downcasting', True)
            MarketingUpdatesMod = MarketingUpdatesMod.replace(0, np.nan)

            UpdatedPayPlans = PayPlans.copy()
            UpdatedPayPlans = pd.DataFrame({
                'Payment Ruleset Code:Name_copy': UpdatedPayPlans['Payment Ruleset Code:Name']
            }).join(UpdatedPayPlans)
            UpdatedPayPlans.set_index('Payment Ruleset Code:Name', inplace=True)

            # amount_columns = UpdatedPayPlans.filter(like='Amount ($)').columns
            # amount_columns = [col[-5:] for col in amount_columns]
            # MarketingUpdatesMod[amount_columns] = MarketingUpdatesMod[amount_columns].astype(float)
            UpdatedPayPlans = UpdatedPayPlans.drop('RULENAME', axis=1)
            UpdatedPayPlans.update(MarketingUpdatesMod, overwrite=True)
            UpdatedPayPlans.update(SpecialtyMod)
            priority_columns = UpdatedPayPlans.filter(like='Priority:').columns
            UpdatedPayPlans[priority_columns] = UpdatedPayPlans[priority_columns].astype('Int64')

            ##############################################
            placeholder = 99999
            differences = OriginalPayPlans.fillna(placeholder).ne(UpdatedPayPlans.fillna(placeholder))
            row_specialty_differences = differences.loc[differences.index.isin(SpecialtyMod.index)]
            SpecialtyColumns = SpecialtyMod.columns.intersection(differences.columns)
            specialty_differences = row_specialty_differences[SpecialtyColumns]

            row_marketing_differences = differences.loc[differences.index.isin(MarketingUpdatesMod.index)]
            MarketingColumns = MarketingUpdatesMod.columns.intersection(differences.columns)
            marketing_differences = row_marketing_differences[MarketingColumns]

            def highlight_cells_combined(x, diff1, diff2, style1, style2):
                """
                Applies and potentially combines distinct styles based on conditions from two difference DataFrames.

                Parameters:
                - x: DataFrame to style.
                - diff1, diff2: Two sets of differences as boolean DataFrames.
                - style1, style2: CSS style strings to apply for diff1 and diff2, respectively.
                """
                styles = pd.DataFrame('', index=x.index, columns=x.columns)  # Initialize with empty styles

                # Function to combine styles, avoiding duplicates
                def combine_styles(style_a, style_b):
                    if style_a and style_b:
                        return style_a + style_b if style_a not in style_b else style_b
                    return style_a or style_b

                # Iterate through each cell to apply styles
                for col in set(diff1.columns).union(diff2.columns).intersection(x.columns):
                    for row in x.index:
                        style_applied = ''
                        if col in diff1.columns and row in diff1.index and diff1.at[row, col]:
                            style_applied = combine_styles(style_applied, style1)
                        if col in diff2.columns and row in diff2.index and diff2.at[row, col]:
                            style_applied = combine_styles(style_applied, style2)
                        styles.at[row, col] = style_applied

                return styles

            # Apply the styling
            PayPlansNew = UpdatedPayPlans.copy()
            # Define your styles for each set of differences
            style1 = 'background-color: #add8e6; color: black'  # For light blue marketing_differences
            style2 = 'background-color: #90EE90; color: black'  # For light green specialty_differences

            # Apply the styling using a lambda function
            UpdatedPayPlans = UpdatedPayPlans.style.apply(
            lambda x: highlight_cells_combined(x, marketing_differences, specialty_differences, style1, style2),
                axis=None)

            ###############################################
            st.dataframe(UpdatedPayPlans)
            # PayPlansNew = st.data_editor(UpdatedPayPlans)
            st.success('Uploaded Data Updated (See highlighted changes)')

        ########################################################################################################################
            def PrepareFile(df, diff1, diff2, file_like_or_path):
                # Save the uploaded file to a temporary file
                if isinstance(file_like_or_path, str):
                    tmp_path = file_like_or_path
                else:
                    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
                        tmp.write(uploaded_PayPlan.getvalue())
                        tmp_path = tmp.name

                # Load the workbook and select the active worksheet
                wb = load_workbook(filename=tmp_path)
                ws = wb.active

                # Assuming your data starts from column A and 'df' is your modified DataFrame
                # Iterate over the DataFrame and update the cells starting from row 13 in Excel (index 12 in openpyxl)
                for r in range(df.shape[0]):
                    for c in range(df.shape[1]):
                        if not pd.isna(df.iloc[r, c]):
                            cell = ws.cell(row=r + 13, column=c + 1, value=df.iloc[r, c])
                            column_name = df.columns[c]
                            row_name = df.index[r]
                            if column_name in diff1.columns and row_name in diff1.index:
                                if diff1.loc[row_name, column_name]:
                                    cell.fill = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')
                                    cell.font = Font(color='000000')
                            if column_name in diff2.columns and row_name in diff2.index:
                                if diff2.loc[row_name, column_name]:
                                    cell.fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
                                    cell.font = Font(color='000000')

                # Save the modified workbook to another temporary file
                with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_modified:
                    wb.save(tmp_modified.name)
                    tmp_modified_path = tmp_modified.name

                # Indicate to the user that the file is ready for download
                st.session_state['file_ready_for_download'] = True

                return tmp_modified_path


            if uploaded_PayPlan is not None:
                tmp_modified_file = PrepareFile(PayPlansNew, marketing_differences, specialty_differences, uploaded_PayPlan)

            # Conditional to check if the file is ready for download
            if 'file_ready_for_download' in st.session_state and st.session_state['file_ready_for_download']:
                # Read the modified file back into memory to serve it via the download button
                with open(tmp_modified_file, 'rb') as file:
                    btn = st.download_button(
                        label="Download Modified Payment Plan File",
                        data=file,
                        file_name="Modified_Payment_Plans.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
